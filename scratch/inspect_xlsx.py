import openpyxl
import os

def inspect_file(filepath):
    print(f"=== Inspecting {os.path.basename(filepath)} ===")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    print("Sheets:", wb.sheetnames)
    for sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        print(f"  Sheet: {sheetname}")
        # Print first 5 rows
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= 5:
                break
            # print non-empty columns
            row_vals = [str(x)[:30] for x in row if x is not None]
            if row_vals:
                print(f"    Row {i+1}: {row_vals[:8]}")

files = [
    r'd:\HOUSEHOLDSURVEY\Staff and Event details - CARE.xlsx',
    r'd:\HOUSEHOLDSURVEY\H.Survey Master File - CARE.xlsx'
]

for f in files:
    if os.path.exists(f):
        inspect_file(f)
    else:
        print(f"File not found: {f}")
